"""app/ui/pages/loader_page.py – Araç Yükleme (QR'li)
=====================================================
• Barkod okutuldukça paket `loaded=1`, `loaded_by`, `loaded_time` güncellenir.
• Liste yalnızca **en az bir paketi yüklenmiş** sevkiyatları gösterir.
• "Liste Yazdır (QR)" butonu: sevkiyat başlığına `qr_token` üretir, QR kodlu PDF oluşturur.
"""
from __future__ import annotations
import csv, os, io, uuid, getpass, sys
from pathlib import Path
from datetime import date, datetime
from typing import Dict, List
from textwrap import wrap

import qrcode
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

from PyQt5.QtCore import Qt, QDate, QTimer, QUrl
from PyQt5.QtGui import QCursor
from PyQt5.QtMultimedia import QSoundEffect
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QDateEdit, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QFileDialog, QMenu, QDialog, QListWidget, QListWidgetItem, QAbstractItemView
)

import app.settings as st
from app import settings, toast
from app.utils.fonts import register_pdf_font
from app.shipment import (
    list_headers_range, trip_by_barkod,
    mark_loaded, set_trip_closed
)
from app.dao.logo import exec_sql, ensure_qr_token, fetch_all, fetch_one

# Use WMS paths instead of relative to file
from app.utils.wms_paths import get_wms_folders
wms_folders = get_wms_folders()
OUTPUT_DIR = wms_folders['output']

# For sounds, use resource path (handles frozen exe)
if getattr(sys, 'frozen', False):
    # In frozen exe, use _MEIPASS for resources
    BASE_DIR = Path(sys._MEIPASS)
else:
    BASE_DIR = Path(__file__).resolve().parents[3]

SOUND_DIR = BASE_DIR / "sounds"

# ═══════════════════════════════════════════════════════════════
# Ses yönetimi - Merkezi sound manager kullan
# ═══════════════════════════════════════════════════════════════
from app.utils.sound_manager import get_sound_manager

# Sound manager instance - memory leak önlenir
sound_manager = get_sound_manager()

# ───────────────────────── Tablo kolonları
COLS = [
    ("id",           "#"),
    ("order_no",     "Sipariş"),
    ("customer_code","Cari Kod"),
    ("customer_name","Müşteri"),
    ("region",       "Bölge"),
    ("address1",     "Adres"),
    ("pkgs_total",   "Paket"),
    ("pkgs_loaded",  "Yüklendi"),
    ("loaded_at",    "Yüklendi 🕒"),
    ("status_txt",   "Durum"),
]


# >>>>> EKLE >>>>>
class ColumnSelectDialog(QDialog):
    """Excel/CSV'de hangi kolonlar olsun?"""
    def __init__(self, parent, cols):
        super().__init__(parent)
        self.setWindowTitle("Kolon Seç")
        self.resize(250, 300)
        v = QVBoxLayout(self)

        self.lst = QListWidget(selectionMode=QAbstractItemView.MultiSelection)
        for key, header in cols:
            itm = QListWidgetItem(header)
            itm.setData(Qt.UserRole, key)
            itm.setSelected(True)           # varsayılan: hepsi
            self.lst.addItem(itm)
        v.addWidget(self.lst)

        btn_ok = QPushButton("Tamam")
        btn_ok.clicked.connect(self.accept)
        v.addWidget(btn_ok, alignment=Qt.AlignRight)

    def selected_keys(self):
        return [i.data(Qt.UserRole) for i in self.lst.selectedItems()]

def _ask_columns(parent) -> list[str] | None:
    dlg = ColumnSelectDialog(parent, COLS)
    return dlg.selected_keys() if dlg.exec_() else None
# <<<<< EKLE <<<<<

# ════════════════════════ UI ═══════════════════════════════════
class LoaderPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()
        # ► Otomatik yenileme – her 30 sn
        self._timer = QTimer(self)
        self._timer.timeout.connect(self.refresh)
        self._timer.start(st.get("loader.auto_refresh", 30) * 1000)         # 30 000 ms = 30 sn
        
        # ═══════════════════════════════════════════════════════════════
        # PERFORMANS İYİLEŞTİRMESİ 2: Barkod okuma optimizasyonu
        # ═══════════════════════════════════════════════════════════════
        self._scanning = False  # Barkod okuma durumu
        
        # ╔════════════════════════════════════════════════════════════╗
        # ║ 🔧 ÇOK SEVİYELİ SIRALAMA SİSTEMİ                         ║
        # ╚════════════════════════════════════════════════════════════╝
        self._sort_history = []  # [(column_index, order), ...]
        self._max_sort_levels = 3  # En fazla 3 seviye sıralama

    def showEvent(self, event):
        """Sekmeye/ekrana dönüldüğünde barkod girişine odaklan."""
        super().showEvent(event)
        
        # ► her gösterimde barkod kutusuna odak
        QTimer.singleShot(0, self.entry.setFocus)
        
        # ► otomatik yenileme yeniden başlasın
        self._timer.start()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.addWidget(QLabel("<b>Araç Yükleme</b>"))

        # — filtre barı —
        top = QHBoxLayout()
        self.dt_from = QDateEdit(QDate.currentDate()); top.addWidget(QLabel("Baş:")); top.addWidget(self.dt_from)
        self.dt_to   = QDateEdit(QDate.currentDate()); top.addWidget(QLabel("Bitiş:")); top.addWidget(self.dt_to)
        self.search  = QLineEdit(); self.search.setPlaceholderText("Ara… (sipariş/cari/bölge)")
        top.addWidget(self.search, 1)
        btn_list   = QPushButton("Yüklemeleri Getir"); btn_list.clicked.connect(self.refresh)
        btn_csv    = QPushButton("Excel/CSV");        btn_csv.clicked.connect(self.export_csv)
        btn_print  = QPushButton("Liste Yazdır (QR)"); btn_print.clicked.connect(self.print_loading_list)  # ★
        btn_done   = QPushButton("Yükleme Tamam");    btn_done.clicked.connect(self.close_trip)
        top.addStretch(); top.addWidget(btn_list); top.addWidget(btn_csv); top.addWidget(btn_print); top.addWidget(btn_done)
        lay.addLayout(top)

        # — tablo —
        self.tbl = QTableWidget(0, len(COLS))
        self.tbl.setHorizontalHeaderLabels([h for _k, h in COLS])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # ╔════════════════════════════════════════════════════════════╗
        # ║ 🔧 ÖZEL SIRALAMA: PyQt5 default sıralama kapalı           ║
        # ╚════════════════════════════════════════════════════════════╝
        self.tbl.setSortingEnabled(False)  # Kendi sıralamamızı kullanacağız
        self.tbl.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tbl.customContextMenuRequested.connect(self._ctx_menu)
        lay.addWidget(self.tbl)

        # — barkod entry —
        bar = QHBoxLayout()
        self.entry = QLineEdit(); self.entry.setPlaceholderText("Paket barkodu → Enter")
        self.entry.returnPressed.connect(self.on_scan); bar.addWidget(self.entry)
        lay.addLayout(bar)

    # ═══════════════════════════════════════════════════════════════
    # PERFORMANS İYİLEŞTİRMESİ 3: Timer optimizasyonu
    # ═══════════════════════════════════════════════════════════════
    def hideEvent(self, event):
        """Sayfa gizlendiğinde timer'ı durdur"""
        super().hideEvent(event)
        self._timer.stop()

    # ══════════════ Veri yükle & tablo doldur ═══════════════════
    def refresh(self):
        # ═══════════════════════════════════════════════════════════════
        # PERFORMANS İYİLEŞTİRMESİ 4: Barkod okuma sırasında refresh'i engelle
        # ═══════════════════════════════════════════════════════════════
        if self._scanning:
            return  # Barkod okuma sırasında refresh yapma
        
        # ╔════════════════════════════════════════════════════════════╗
        # ║ 🎯 FIX: Seçili satırları koru (trip_id bazında)           ║
        # ╚════════════════════════════════════════════════════════════╝
        selected_trip_ids = []
        if self.tbl.rowCount() > 0:
            # Fix: PyQt5'te doğru seçim kontrolü
            selected_rows = [index.row() for index in self.tbl.selectionModel().selectedRows()]
            for row_idx in selected_rows:
                if self.tbl.item(row_idx, 0):
                    trip_id = int(self.tbl.item(row_idx, 0).text())
                    selected_trip_ids.append(trip_id)
            
        d1 = self.dt_from.date().toPyDate().isoformat()
        d2 = self.dt_to.date().toPyDate().isoformat()
        rows = list_headers_range(d1, d2)

        # Arama filtresi
        q = self.search.text().strip().upper()
        if q:
            rows = [r for r in rows if q in r["order_no"].upper()
                               or q in (r["customer_code"] or "").upper()
                               or q in (r["region"] or "").upper()]

        # Yalnızca en az 1 paket yüklenmişse göster
        rows = [r for r in rows if r["pkgs_loaded"] > 0]

        # Başlık satırı ikon-metni
        for r in rows:
            r["status_txt"] = (
                "🚚" if r.get("en_route")                      # araç yolda
                else "✔" if r["closed"]                       # tamamen yüklü
                else "⏳"                                      # bekliyor
            )
            r["loaded_at"] = (r.get("loaded_at") or "")[:19]

        # Tabloyu güncelle
        self._rows   = rows
        self._id_map = {r["id"]: r for r in rows}             # 🔸 fix: tüm id'ler
        
        # ╔════════════════════════════════════════════════════════════╗
        # ║ 🎯 FIX: Mevcut sıralamayı koru                            ║
        # ╚════════════════════════════════════════════════════════════╝
        if self._sort_history:
            # Eğer sıralama varsa uygula
            self._apply_multi_sort()
        else:
            # Sıralama yoksa normal yenileme
            self.tbl.setRowCount(0)
            for rec in rows:
                self._add_row(rec)
        
        # ╔════════════════════════════════════════════════════════════╗
        # ║ 🔄 FIX: Seçimleri geri yükle                              ║
        # ╚════════════════════════════════════════════════════════════╝
        if selected_trip_ids:
            for row_idx in range(self.tbl.rowCount()):
                if self.tbl.item(row_idx, 0):
                    trip_id = int(self.tbl.item(row_idx, 0).text())
                    if trip_id in selected_trip_ids:
                        self.tbl.selectRow(row_idx)
        
        # Focus: Eğer seçim yoksa barkod kutusuna, varsa tablo seçimini koru
        if not selected_trip_ids:
            self.entry.setFocus(Qt.OtherFocusReason)
    # ────────────────────────────────────────────────────────────

    # ╔════════════════════════════════════════════════════════════╗
    # ║ 🎯 ÇOK SEVİYELİ SIRALAMA SİSTEMİ                         ║
    # ╚════════════════════════════════════════════════════════════╝
    def _on_header_clicked(self, column_index: int):
        """Header'a tıklandığında çok seviyeli sıralama yap"""
        from PyQt5.QtCore import Qt as QtCore
        
        # Mevcut sıralama durumunu belirle
        current_order = QtCore.AscendingOrder
        
        # Eğer bu kolon zaten sıralama geçmişinde varsa, sırayı ters çevir
        for i, (col, order) in enumerate(self._sort_history):
            if col == column_index:
                current_order = QtCore.DescendingOrder if order == QtCore.AscendingOrder else QtCore.AscendingOrder
                # Bu kolonu geçmişten kaldır (en üstte olacak)
                self._sort_history.pop(i)
                break
        
        # Bu kolonu geçmişin en başına ekle
        self._sort_history.insert(0, (column_index, current_order))
        
        # Maksimum seviye kontrolü
        if len(self._sort_history) > self._max_sort_levels:
            self._sort_history = self._sort_history[:self._max_sort_levels]
        
        # Sıralamayı uygula
        self._apply_multi_sort()
        
        # Header'da sıralama göstergesi göster
        self._update_header_indicators()
    
    def _apply_multi_sort(self):
        """Çok seviyeli sıralama uygula"""
        if not hasattr(self, '_rows') or not self._rows:
            return
        
        from PyQt5.QtCore import Qt as QtCore
        
        def sort_key(row):
            """Sıralama anahtarı oluştur"""
            keys = []
            for col_idx, order in self._sort_history:
                # Kolon anahtarını al
                col_key = COLS[col_idx][0]
                value = row.get(col_key, "")
                
                # Sayısal değerler için özel işleme
                if col_key in ("id", "pkgs_total", "pkgs_loaded"):
                    try:
                        value = int(value) if value else 0
                    except (ValueError, TypeError):
                        value = 0
                
                # Tarih değerleri için özel işleme
                elif col_key in ("loaded_at", "created_at"):
                    if not value or value == "":
                        value = "1900-01-01"  # En eski tarih
                
                # String değerler için büyük/küçük harf duyarsız
                else:
                    value = str(value).upper()
                
                # Ters sıralama için değeri ters çevir
                if order == QtCore.DescendingOrder:
                    if isinstance(value, (int, float)):
                        value = -value
                    else:
                        # String için reverse mantığı
                        value = f"zzzzz_{value}"  # Alfabetik tersleme
                
                keys.append(value)
            
            return keys
        
        # Sırala
        sorted_rows = sorted(self._rows, key=sort_key)
        
        # Tabloyu güncelle
        self._rows = sorted_rows
        self._id_map = {r["id"]: r for r in sorted_rows}
        
        # Seçili satırları koru
        selected_trip_ids = []
        if self.tbl.rowCount() > 0:
            selected_rows = [index.row() for index in self.tbl.selectionModel().selectedRows()]
            for row_idx in selected_rows:
                if self.tbl.item(row_idx, 0):
                    trip_id = int(self.tbl.item(row_idx, 0).text())
                    selected_trip_ids.append(trip_id)
        
        # Tabloyu yeniden doldur
        self.tbl.setRowCount(0)
        for rec in sorted_rows:
            self._add_row(rec)
        
        # Seçimleri geri yükle
        if selected_trip_ids:
            for row_idx in range(self.tbl.rowCount()):
                if self.tbl.item(row_idx, 0):
                    trip_id = int(self.tbl.item(row_idx, 0).text())
                    if trip_id in selected_trip_ids:
                        self.tbl.selectRow(row_idx)
    
    def _update_header_indicators(self):
        """Header'larda sıralama göstergelerini güncelle"""
        from PyQt5.QtCore import Qt as QtCore
        
        # Tüm header'ları temizle
        for i in range(len(COLS)):
            header_text = COLS[i][1]  # Orijinal header text
            self.tbl.setHorizontalHeaderItem(i, QTableWidgetItem(header_text))
        
        # Sıralama göstergelerini ekle
        for priority, (col_idx, order) in enumerate(self._sort_history):
            current_text = COLS[col_idx][1]
            
            # Sıralama yönü göstergesi
            direction = "↑" if order == QtCore.AscendingOrder else "↓"
            
            # Öncelik göstergesi (1, 2, 3...)
            priority_indicator = f"({priority + 1})" if len(self._sort_history) > 1 else ""
            
            # Yeni header text
            new_text = f"{current_text} {direction}{priority_indicator}"
            self.tbl.setHorizontalHeaderItem(col_idx, QTableWidgetItem(new_text))
    
    def clear_sorting(self):
        """Tüm sıralamaları temizle (Ctrl+Click için)"""
        self._sort_history.clear()
        self._update_header_indicators()
        # Orijinal sıralamaya geri dön (ID'ye göre)
        if hasattr(self, '_rows') and self._rows:
            self._rows.sort(key=lambda x: x.get("id", 0))
            self._id_map = {r["id"]: r for r in self._rows}
            
            # Tabloyu yeniden doldur
            self.tbl.setRowCount(0)
            for rec in self._rows:
                self._add_row(rec)


    def _add_row(self, rec: Dict):
        r = self.tbl.rowCount(); self.tbl.insertRow(r)
        for c, (k, _h) in enumerate(COLS):
            itm = QTableWidgetItem(str(rec.get(k, "")))
            itm.setTextAlignment(Qt.AlignCenter)
            # renk mantığı - FIX: Sarı (kısmi yükleme) eklendi
            if rec["pkgs_loaded"] >= rec["pkgs_total"]:
                itm.setBackground(Qt.green)
            elif rec["pkgs_loaded"] == 0:
                itm.setBackground(Qt.red)
            else:
                # Kısmi yükleme - sarı arka plan
                from PyQt5.QtGui import QColor
                itm.setBackground(QColor(255, 255, 0))
            self.tbl.setItem(r, c, itm)

    # ═══════════════════════════════════════════════════════════════
    # PERFORMANS İYİLEŞTİRMESİ 5: Tek satır güncelleme metodu
    # ═══════════════════════════════════════════════════════════════
    def _update_single_trip(self, trip_id: int):
        """Sadece belirtilen trip_id'nin satırını güncelle - FULL REFRESH YOK"""
        try:
            # 1. Güncel veriyi al
            updated_row = fetch_one("""
                SELECT h.id, h.order_no, h.customer_code, h.customer_name, h.region, 
                       h.address1, h.pkgs_total, h.closed, h.created_at, h.en_route,
                       COALESCE(l.loaded_count, 0) as pkgs_loaded,
                       l.loaded_at
                FROM shipment_header h
                LEFT JOIN (
                    SELECT trip_id, COUNT(*) as loaded_count, MAX(loaded_at) as loaded_at
                    FROM shipment_loaded 
                    WHERE trip_id = ? AND loaded = 1
                    GROUP BY trip_id
                ) l ON h.id = l.trip_id
                WHERE h.id = ?
            """, trip_id, trip_id)
            
            if not updated_row:
                return
                
            # Status text ekle
            updated_row["status_txt"] = (
                "🚚" if updated_row.get("en_route") 
                else "✔" if updated_row["closed"] 
                else "⏳"
            )
            updated_row["loaded_at"] = (updated_row.get("loaded_at") or "")[:19]
            
            # 2. Tabloda satırı bul ve güncelle
            for row_idx in range(self.tbl.rowCount()):
                if int(self.tbl.item(row_idx, 0).text()) == trip_id:
                    # Satırı güncelle
                    for c, (k, _h) in enumerate(COLS):
                        item = self.tbl.item(row_idx, c)
                        new_value = str(updated_row.get(k, ""))
                        if item.text() != new_value:
                            item.setText(new_value)
                            # Renk güncelle
                            if updated_row["pkgs_loaded"] >= updated_row["pkgs_total"]:
                                item.setBackground(Qt.green)
                            elif updated_row["pkgs_loaded"] == 0:
                                item.setBackground(Qt.red)
                            else:
                                from PyQt5.QtGui import QColor
                                item.setBackground(QColor(255, 255, 0))  # Partial loading - yellow
                    break
            
            # 3. Internal cache'i güncelle
            if hasattr(self, '_id_map'):
                self._id_map[trip_id] = updated_row
            if hasattr(self, '_rows'):
                for i, row in enumerate(self._rows):
                    if row["id"] == trip_id:
                        self._rows[i] = updated_row
                        break
                        
        except Exception as e:
            # ⚠️ İyileştirilmiş hata yönetimi
            error_msg = f"Trip güncelleme hatası (ID: {trip_id})"
            print(f"❌ {error_msg}: {e}")
            toast("Güncelleme Hatası", "Veriler yeniden yüklenecek...")
            # Fallback: 2 saniye sonra full refresh
            QTimer.singleShot(2000, self.refresh)

    # ══════════════ Barkod okuma ════════════════════════════════
    def on_scan(self):
        # ═══════════════════════════════════════════════════════════════
        # PERFORMANS İYİLEŞTİRMESİ 6: Barkod okuma süresince flag set et
        # ═══════════════════════════════════════════════════════════════
        if self._scanning:
            return  # Önceki scan henüz bitmemiş
            
        self._scanning = True  # Flag set
        
        try:
            raw = self.entry.text().strip()
            self.entry.clear()
            if not raw or "-K" not in raw:
                sound_manager.play_error()                      # 🔊 hata
                return

            inv_root, pkg_txt = raw.rsplit("-K", 1)
            try:
                pkg_no = int(pkg_txt)
            except ValueError:
                sound_manager.play_error()                      # 🔊 hata
                return

            # ► Aktif sevkiyat başlığını bul
            trip = trip_by_barkod(inv_root)          # tarih filtresiz
            if not trip:
                sound_manager.play_error()                      # 🔊 hata
                QMessageBox.warning(self, "Paket", "Sevkiyat başlığı bulunamadı!")
                return

            trip_id, pkg_tot = trip
            
            # ► Güncel pkgs_total değerini al (backorder etiket güncellemesi için)
            current_header = fetch_one(
                "SELECT pkgs_total FROM shipment_header WHERE id = ?", 
                trip_id
            )
            if current_header:
                pkg_tot = current_header["pkgs_total"]
            
            # ► Paket sayısı azaltıldıysa kontrol et
            if pkg_no > pkg_tot:
                # Paketi kontrol et - yüklenmiş mi?
                loaded_check = fetch_one(
                    "SELECT loaded FROM shipment_loaded WHERE trip_id = ? AND pkg_no = ?",
                    trip_id, pkg_no
                )
                
                if loaded_check and loaded_check["loaded"] == 1:
                    # Yüklenmiş paket, silinemez!
                    sound_manager.play_error()                      # 🔊 hata
                    QMessageBox.critical(self, "Kritik Hata", 
                        f"Paket #{pkg_no} zaten yüklenmiş durumda!\n"
                        f"Yüklenmiş paketler silinemez.\n"
                        f"Paket sayısı en az {pkg_no} olmalıdır.")
                    return
                else:
                    # Yüklenmemiş fazla paket, güvenli atomic delete işlemi
                    try:
                        from app.dao.transactions import transaction_scope
                        
                        # Atomic transaction kullanarak güvenli delete
                        with transaction_scope() as conn:
                            cursor = conn.cursor()
                            
                            # Triple check with row locking - race condition koruması
                            cursor.execute(
                                """
                                SELECT loaded FROM shipment_loaded WITH (UPDLOCK, ROWLOCK)
                                WHERE trip_id = ? AND pkg_no = ?
                                """,
                                trip_id, pkg_no
                            )
                            final_check = cursor.fetchone()
                            
                            if final_check and final_check[0] == 0:
                                # Hala yüklenmemiş ve lock'lanmış, güvenle sil
                                cursor.execute(
                                    """
                                    DELETE FROM shipment_loaded 
                                    WHERE trip_id = ? AND pkg_no = ? AND loaded = 0
                                    """,
                                    trip_id, pkg_no
                                )
                                
                                # Verify deletion was successful
                                if cursor.rowcount > 0:
                                    sound_manager.play_error()
                                    QMessageBox.warning(self, "Paket", 
                                        f"Paket numarası geçersiz! (1-{pkg_tot} arası olmalı)\n"
                                        f"Fazla paket kaydı güvenli şekilde silindi.")
                                else:
                                    sound_manager.play_error()
                                    QMessageBox.critical(self, "Kritik Hata", 
                                        f"Paket #{pkg_no} silinirken beklenmeyen durum!\n"
                                        f"İşlem iptal edildi.")
                            else:
                                # Bu arada yüklenmiş olabilir
                                sound_manager.play_error()
                                QMessageBox.critical(self, "Kritik Hata", 
                                    f"Paket #{pkg_no} silme işlemi sırasında yüklenmiş!\n"
                                    f"İşlem iptal edildi. Yüklenmiş paketler silinemez.")
                                    
                    except Exception as e:
                        sound_manager.play_error()
                        QMessageBox.critical(self, "Database Hatası", 
                            f"Paket silme işlemi başarısız: {str(e)}")
                        
                    return
            
            if not (1 <= pkg_no <= pkg_tot):
                sound_manager.play_error()                      # 🔊 hata
                QMessageBox.warning(self, "Paket", f"Paket numarası geçersiz! (1-{pkg_tot} arası olmalı)")
                return

            # ──────────────────────────────────────────────
            # 1) shipment_loaded + shipment_header güncelle
            # ──────────────────────────────────────────────
            ok = mark_loaded(trip_id, pkg_no)
            if ok == 0:                 # yinelenen okuma
                sound_manager.play_duplicate()               # 🔊 tekrar
                toast("Uyarı", "Bu paket zaten yüklenmiş!")
                return

            # ──────────────────────────────────────────────
            # 2) İlgili shipment_lines satırlarını işaretle
            # ──────────────────────────────────────────────
            hdr = fetch_one(
                "SELECT order_no, trip_date "
                "FROM   shipment_header "
                "WHERE  id = ?", trip_id
            )
            if hdr:
                try:
                    # UPDATE işlemi
                    exec_sql(
                        """
                        UPDATE shipment_lines
                           SET loaded = 1
                         WHERE order_no  = ?
                           AND trip_date = ?
                           AND loaded = 0;""",
                        hdr["order_no"], hdr["trip_date"]
                    )
                except Exception as e:
                    print(f"❌ shipment_lines güncelleme hatası: {e}")
                    # Hata olsa bile devam et

            sound_manager.play_ok()                         # 🔊 başarılı okuma
            toast("Paket Yüklendi", f"{inv_root} K{pkg_no}")
            
            # ╔════════════════════════════════════════════════════════════╗
            # ║ 🚀 PERFORMANS İYİLEŞTİRMESİ: Tek satır güncelleme         ║
            # ║ Full refresh yerine sadece etkilenen trip güncellenir     ║
            # ╚════════════════════════════════════════════════════════════╝
            self._update_single_trip(trip_id)
            
            # Focus geri ver - kullanıcı deneyimi için kritik
            QTimer.singleShot(100, self.entry.setFocus)
            
        finally:
            self._scanning = False  # Flag temizle

    # ════════════════════════════════════════════════════════════
    # ───────────── Uygulama Ayarları Anında Uygula ─────────────
    def apply_settings(self):
        """MainWindow -> _apply_global_settings çağırır."""
        # ► Otomatik yenile
        self._timer.setInterval(st.get("loader.auto_refresh", 30) * 1000)

        # ► Otomatik fokus
        self._auto_focus = st.get("ui.auto_focus", True)

        # ► Ses
        # Sound manager ayarlarını uygula
        sound_manager.apply_settings()



        # ════════════════════════════════════════════════════════════
        # ════════════════════════════════════════════════════════════
    def print_loading_list(self):
        if not getattr(self, "_rows", None):
            QMessageBox.warning(self, "Liste", "Önce listeyi getir!")
            return

        visible_ids = [
            int(self.tbl.item(r, 0).text()) for r in range(self.tbl.rowCount())
        ]
        rows_in_view = [self._id_map[i] for i in visible_ids]

        sel_rows = {ix.row() for ix in self.tbl.selectionModel().selectedRows()}
        rows_to_print = [rows_in_view[r] for r in sel_rows] if sel_rows else rows_in_view
        if not rows_to_print:
            QMessageBox.information(self, "Liste", "Basılacak satır yok.")
            return

        out_pdf = OUTPUT_DIR / f"loader_{datetime.now():%Y%m%d_%H%M%S}.pdf"

        FONT = register_pdf_font()  # ← merkezi font yönetimi
        W, H = landscape(A4)
        pdf = canvas.Canvas(str(out_pdf), pagesize=(W, H))
        pdf.setFont(FONT, 8)

        cols = [
            ("QR",        22*mm), ("Sipariş",   28*mm),
            ("Cari Kod",  24*mm), ("Müşteri",   38*mm),
            ("Bölge",     28*mm), ("Adres",     50*mm),
            ("Paket",     10*mm), ("Yüklendi",  32*mm),
            ("Kaşe",      40*mm),
        ]
        margin, header_h, row_h_min = 15*mm, 12*mm, 24*mm
        y_top = H - margin
        total_pkgs = sum(r["pkgs_total"] for r in rows_to_print)

        def split_text(txt, font, size, max_w):
            out, cur = [], ""
            for w in str(txt).split():
                test = (cur + " " + w).strip()
                if stringWidth(test, font, size) <= max_w:
                    cur = test
                else:
                    if cur: out.append(cur); cur = w
            out.append(cur); return out

        def draw_header(y):
            pdf.setFont(FONT, 10)
            pdf.drawString(margin, y + 4*mm, f"Tarih: {date.today():%d.%m.%Y}    Toplam Koli: {total_pkgs}")
            pdf.setFont(FONT, 8)
            x = margin
            for title, w in cols:
                pdf.rect(x, y-header_h, w, header_h)
                pdf.drawCentredString(x + w/2, y-header_h + 3, title)
                x += w

        draw_header(y_top); y_cursor = y_top - header_h

        for rec in rows_to_print:
            buf = io.BytesIO()
            qrcode.make(ensure_qr_token(rec["order_no"])).save(buf, "PNG")
            qr_img = ImageReader(buf); buf.seek(0)

            cell_vals = [
                rec["order_no"], rec["customer_code"], rec["customer_name"],
                rec["region"], rec["address1"],
                f"{rec['pkgs_loaded']} / {rec['pkgs_total']}",
                rec["loaded_at"][:19], "",
            ]

            dyn_row_h, cell_lines = row_h_min, []
            for (_t, w), txt in zip(cols[1:], cell_vals):
                lines = split_text(txt, FONT, 7, w-4*mm)
                cell_lines.append(lines)
                dyn_row_h = max(dyn_row_h, 6 + 9*len(lines))

            if y_cursor - dyn_row_h < margin:
                pdf.showPage(); pdf.setFont(FONT, 8)
                draw_header(H - margin)
                y_cursor = H - margin - header_h

            x = margin
            for _t, w in cols:
                pdf.rect(x, y_cursor-dyn_row_h, w, dyn_row_h)
                x += w

            qr_sz = 18*mm
            pdf.drawImage(
                qr_img,
                margin + (cols[0][1]-qr_sz)/2,
                y_cursor - dyn_row_h + (dyn_row_h-qr_sz)/2,
                qr_sz, qr_sz, preserveAspectRatio=True
            )

            x = margin + cols[0][1]
            pdf.setFont(FONT, 7)
            for (_t, w), lines in zip(cols[1:], cell_lines):
                for i, line in enumerate(lines):
                    pdf.drawString(x+2, y_cursor - 9 - i*9, line)
                x += w

            y_cursor -= dyn_row_h

        pdf.save()
        os.startfile(out_pdf)
        toast("PDF Hazır", str(out_pdf))




    def split_text(text: str, font_name: str, font_size: int, max_width: float):
        """
        max_width (pt) değerini aşmadan kelimeleri satırlara ayır.
        """
        words  = text.split()
        lines  = []
        line   = ""
        for w in words:
            test = (line + " " + w).strip()
            if stringWidth(test, font_name, font_size) <= max_width:
                line = test
            else:
                if line:        # önceki satırı kaydet
                    lines.append(line)
                line = w        # kelimeyi yeni satıra taşı
        lines.append(line)
        return lines
    # ══════════════ Manuel kapama ═══════════════════════════════
    def close_trip(self):
        """
        Seçili sevkiyat(lar)ı kapatır.
        • Eksik koli varsa önce onay ister.
        • Eksik kapatma USER_ACTIVITY tablosuna loglanır.
        """
        rows = {i.row() for i in self.tbl.selectedIndexes()}
        if not rows:
            return

        for row in rows:
            trip_id = int(self.tbl.item(row, 0).text())
            rec = self._id_map.get(trip_id)
            if not rec:
                continue

            # Eksik koli var mı?
            if rec["pkgs_loaded"] < rec["pkgs_total"]:
                ans = QMessageBox.question(
                    self, "Eksik Koli",
                    f"{rec['pkgs_loaded']} / {rec['pkgs_total']} yüklendi.\n"
                    "Yine de 'Yükleme Tamam' yapılsın mı?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if ans == QMessageBox.No:
                    continue  # kullanıcı vazgeçti

                # Log – eksik kapatma
                exec_sql("""
                    INSERT INTO USER_ACTIVITY
                        (username, action, details, order_no)
                    SELECT ?, 'TRIP_MANUAL_CLOSED_INCOMPLETE', ?, order_no
                      FROM shipment_header
                     WHERE id = ?""",
                    getpass.getuser(),
                    f"{rec['pkgs_loaded']}/{rec['pkgs_total']}",
                    trip_id
                )

            # Kapama işlemi (sadece en_route = 1, closed = 0 kalsın)
            set_trip_closed(trip_id, closed=False, en_route_only=True)

        self.refresh()


   # ─────────────────────────────────────────────────────────────
#  Dışa Aktarım  –  CSV / Excel  (kolon seçmeli)
# ─────────────────────────────────────────────────────────────
    def export_csv(self):
        """Mevcut satırları CSV / Excel'e dışa aktarır.
        • Önce kolon seçimi diyalogu açılır.
        • Seçim yapılmazsa (İptal) işlem durur.
        """
        if not getattr(self, "_rows", None):
            QMessageBox.warning(self, "Dışa Aktarım", "Önce listeyi getir!"); return

        sel_keys = _ask_columns(self)                 # ← yeni diyalog
        if not sel_keys:                              # İptal
            return

        fn, _ = QFileDialog.getSaveFileName(
            self, "Kaydet", str(BASE_DIR / f"loader_{date.today():%Y%m%d}"),
            "Excel (*.xlsx);;CSV (*.csv)"
        )
        if not fn:
            return

        if fn.lower().endswith(".csv"):
            self._write_csv(fn, sel_keys)
        else:
            self._write_xlsx(fn, sel_keys)

        QMessageBox.information(self, "Dışa Aktarım", f"Dosya yazıldı:\n{fn}")

        # ---------------- CSV -------------------------------------
    def _write_csv(self, path: str, keys: list[str]):
        """
        Seçili kolonları ('keys') kullanarak CSV oluşturur ve
        tamamlandığında varsayılan programla dosyayı açar.
        """
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([header for k, header in COLS if k in keys])   # başlık
            for rec in self._rows:
                w.writerow([rec.get(k, "") for k, _h in COLS if k in keys])

        os.startfile(path)   # ↻  otomatik aç

    # ---------------- XLSX ------------------------------------
    def _write_xlsx(self, path: str, keys: list[str]):
        """
        Seçili kolonlarla Excel (.xlsx) üretir; bittiğinde otomatik açar.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(self, "Excel", "pip install openpyxl")
            return

        wb = Workbook(); ws = wb.active

        ws.append([header for k, header in COLS if k in keys])        # başlık
        for rec in self._rows:                                        # satırlar
            ws.append([rec.get(k, "") for k, _h in COLS if k in keys])

        # Otomatik sütun genişliği
        for col_idx in range(1, len(keys) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value) or "")
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        wb.save(path)
        os.startfile(path)   # ↻  otomatik aç

    # ══════════════ Sağ‑tık Detay ═══════════════════════════════
    def _ctx_menu(self, pos):
        idx = self.tbl.indexAt(pos); row = idx.row()
        if row < 0:
            return
        trip_id = int(self.tbl.item(row, 0).text())
        rec = self._id_map.get(trip_id)
        if not rec:
            return
        txt = [f"<b>Sipariş No</b>: {rec['order_no']}"]
        for k in ("customer_code", "customer_name", "region", "address1",
                  "pkgs_total", "pkgs_loaded", "loaded_at", "closed", "created_at"):
            txt.append(f"{k.replace('_',' ').title()}: {rec.get(k, '')}")
        QMessageBox.information(self, "Sipariş Detay", "<br>".join(txt))